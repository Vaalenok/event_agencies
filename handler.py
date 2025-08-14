from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

from models import Company, Person


EMAIL_REGEX = r'[A-Za-z][A-Za-z0-9._%+-]*@[A-Za-z0-9-]+(?:\.[A-Za-z0-9-]+)*\.[A-Za-z]{2,}(?=$|\s|[^\w.-])'
PHONE_REGEX = r'(?<!\d)(?:\+1\s?)?(?:\(\d{3}\)|\d{3})[\s.-]?\d{3}[\s.-]?\d{4}(?![A-Za-z0-9@])'


def find_companies(_json: dict):
    result = []

    vendors = _json.get("vendors", [])

    if vendors:
        for vendor in vendors:
            result.append(
                Company(
                    vendor.get("name", None),
                    None,
                    None,
                    None,
                    vendor.get("phoneNumber", None),
                    [],
                    vendor.get("slug", None)
                )
            )

    return result


def get_company_info(html, company: Company):
    soup = BeautifulSoup(html, "html.parser")

    script_tag = soup.find(
        "script", attrs={"type": "application/ld+json", "data-sentry-component": "StructuredData"}
    )

    if script_tag:
        json_text = script_tag.string
        data = json.loads(json_text)
        website_url = data.get("url", None)

        if website_url:
            if website_url.endswith("/"):
                website_url = website_url[:-1]

        company.website = website_url

    div = soup.find("div", attrs={"data-sentry-component": "SocialLinks"})

    if div:
        links = div.find_all("a")

        if links:
            for link in links:
                url = link.get("href", None)

                if "instagram" in url:
                    company.ig_link = url

                if "facebook" in url:
                    company.fb_link = url

    return company


def get_persons_info(html, company: Company):
    soup = BeautifulSoup(html, "html.parser")
    script_tag = soup.find("script", text=re.compile(r'^self\.__next_f\.push\(\[1,\s*"4'))

    if script_tag:
        match = re.search(r'self\.__next_f\.push\((.*)\)\s*;?', script_tag.string, re.S)

        if match:
            raw = match.group(1)
            arr = json.loads(raw)
            inner_str = arr[1]
            json_match = re.search(r'\{.*\}', inner_str, re.S)

            if json_match:
                json_text = json_match.group(0)
                data = json.loads(json_text)

                raw_persons = data.get("pro", {}).get("teamMembers", [])

                for person in raw_persons:
                    company.persons.append(
                        Person(
                            person.get("name", None),
                            person.get("title", None),
                            None,
                            None
                        )
                    )

    return company


async def get_persons_links(company: Company):
    from main import parser

    visited = set()
    pages_to_visit = [company.website]

    tags = ["about", "contact", "contacts", "team", "us", "meet", "franchise", "welcome"]

    depth = 15

    while pages_to_visit and depth:
        current_page = pages_to_visit.pop(0)

        if current_page not in visited:
            visited.add(current_page)

            website = await parser.parse_website(current_page)

            if website:
                soup = BeautifulSoup(website, "html.parser")

                all_links = soup.find_all("a")

                for link in all_links:
                    link_text = link.get_text(strip=True).lower()
                    href_text = link.get("href", None)

                    if any(tag in link_text for tag in tags) or any(tag in href_text for tag in tags if href_text):
                        if href_text:
                            if href_text.startswith("http"):
                                new_page = href_text
                            elif href_text.startswith("/"):
                                new_page = company.website.rstrip("/") + href_text
                            else:
                                continue

                            if new_page not in visited and new_page not in pages_to_visit:
                                pages_to_visit.append(new_page)

        depth -= 1

    pages = list(visited)

    for page in pages:
        html = await parser.parse_website(page)

        if html:
            soup = BeautifulSoup(html, "html.parser")
            page_text = soup.get_text()

            for person in company.persons:
                match_emails = [[], []]

                found_emails = []
                found_phones = []

                all_emails = re.findall(EMAIL_REGEX, page_text)
                all_phones = re.findall(PHONE_REGEX, page_text)

                first_name = person.name.split(" ")[0].lower()

                if len(person.name.split(" ")) > 1:
                    last_name = person.name.split(" ")[1].lower()
                else:
                    last_name = " "

                for a in soup.find_all("a", href=True):
                    href = a.get("href", None)

                    if href:
                        if href.startswith("mailto:"):
                            email = href[len("mailto:"):]
                            all_emails.append(email)

                for email in all_emails:
                    if not last_name or len(last_name) > 1 and "." not in last_name:
                        if email and (first_name in email.lower() or last_name in email.lower()):
                            found_emails.append(email)

                for phone in all_phones:
                    found_phones.append(phone)

                unique_emails = list(set(found_emails))
                unique_phones = list(set(found_phones))

                if unique_emails:
                    for email in unique_emails:
                        emails = [_person.email for _person in company.persons]
                        email_parts = email.split(".")

                        if "com" in email_parts[-1] and len(email_parts[-1]) != 3:
                            email_parts[-1] = "com"

                        if "area" in email_parts:
                            email_parts.remove("area")

                        new_email = ".".join(email_parts)

                        if new_email not in emails:
                            match_count = -1

                            if first_name in new_email.lower():
                                match_count += 1

                            if last_name in new_email.lower():
                                match_count += 1

                            if match_count > -1:
                                match_emails[match_count].append(new_email)

                if unique_phones:
                    numbers = [_person.phone_number for _person in company.persons]

                    if unique_phones[0] not in numbers:
                        person.phone_number = unique_phones[0]

                if match_emails[1]:
                    person.email = match_emails[1][0]
                elif match_emails[0]:
                    person.email = match_emails[0][0]

    return company


def write_in_xlsx(companies: list[Company]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.sheet_view.showGridLines = False

    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    headers = [
        "Company name", "Company website", "Company instagram link", "Company facebook link", "Company phone number",
        "Person name", "Person job title", "Person phone number", "Person email"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = grey_fill

    current_row = 2

    for idx, company in enumerate(companies):
        use_grey = (idx % 2 == 1)

        rows_to_add = []

        if company.persons:
            for person in company.persons:
                rows_to_add.append([
                    company.name, company.website, company.ig_link, company.fb_link, company.phone_number,
                    person.name, person.job_title, person.phone_number, person.email
                ])
        else:
            rows_to_add.append([
                company.name, company.website, company.ig_link, company.fb_link, company.phone_number
            ])

        for row_data in rows_to_add:
            ws.append(row_data)

            if use_grey:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=current_row, column=col).fill = grey_fill

            current_row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))

                if cell_length > max_length:
                    max_length = cell_length

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("data.xlsx")
